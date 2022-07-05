import csv
import json
import shutil
import time

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from requests import get
from selenium import webdriver

from product import Product


class VW_LCV:
    page_address = "https://vwdostawcze-sklep.pl"

    def __init__(self):
        self.final_data = []
        self.webdriver = "chromedriver.exe"
        product_category = self.get_product_category()
        page = get(self.page_address)
        self.bs = BeautifulSoup(page.content, "html.parser")
        self.set_spider(self.bs, product_category)

    def get_product_category(self):
        wb = load_workbook("products.xlsx")
        ws = wb.worksheets[0]
        return [row for row in ws.iter_rows(values_only=True)]

    def set_spider(self, bs, product_category):
        categories = self.get_category_list(bs)
        self.get_subcategoires(bs, categories)
        products = self.set_products(product_category)
        products = self.get_tshirt_size(products, product_category)
        uni_products = self.remove_duplicate(products)
        self.set_json(uni_products)
        self.set_csv_fb(uni_products)
        self.set_csv_gmc(uni_products)
        print("############ DONE ############")

    def get_category_list(self, bs):
        category_links = bs.find("ul", class_="category-list swiper-wrapper").find_all(
            "a"
        )
        categories = []
        for category_link in category_links:
            categories.append(
                {
                    "link": category_link["href"],
                    "name": category_link.get_text("|", strip=True),
                }
            )
        return categories

    def get_subcategoires(self, bs, categories):
        """
        Get subproducts for each category"
        """
        for category in categories:
            #TODO REMOVE IF
            # if category["name"] == "KOLEKCJA T1":
            subproduct_address = self.page_address + category["link"]
            subproducts_page = get(subproduct_address)
            bs_subproducts_list = BeautifulSoup(
                subproducts_page.content, "html.parser"
            )
            self.final_data.extend(
                self.get_products(
                    bs_subproducts_list, category["name"], subproduct_address
                )
            )

    def get_products(self, subproducts, name, subproduct_addresss):

        finish = True
        products = []
        while finish:
            if len(products) == 0:
                # make new dictinary
                # alists = subproducts.find_all("a", class_="product-list__item")
                subproducts_test = self.scroll_page(subproduct_addresss)
                alists = subproducts_test.find_all("a", class_="product-list__item")

                if len(alists) == 0:
                    continue
                price = (
                    True
                    if alists[0].find("div", class_="product-list__price")
                    else False
                )
                if price:
                    # "no more subcategories"
                    for alist in alists:
                        products.append(
                            {
                                "product_name": name,
                                "products": [f'{alist.find("p").get_text()}'],
                                "path": alist["href"],
                                "finish": True,
                            }
                        )

                else:
                    for alist in alists:
                        products.append(
                            {
                                "product_name": name,
                                "products": [f'{alist.find("p").get_text()}'],
                                "path": alist["href"],
                                "finish": False,
                            }
                        )
            else:
                # add product to key products
                temp_products = []
                for product in products:
                    # path = get(self.page_address + product["path"])
                    # products_bs = BeautifulSoup(path.content, "html.parser")
                    products_bs = self.scroll_page(self.page_address + product["path"])
                    alists = products_bs.find_all("a", class_="product-list__item")
                    if len(alists) == 0:
                        continue
                    price = (
                        True
                        if alists[0].find("div", class_="product-list__price")
                        else False
                    )

                    if price:
                        # "no more subcategories"
                        for alist in alists:
                            temp_product = {}
                            products_list = product["products"].copy()
                            products_list.append(f'{alist.find("p").get_text()}')
                            temp_product["product_name"] = product["product_name"]
                            temp_product["products"] = products_list
                            temp_product["path"] = alist["href"]
                            temp_product["finish"] = True
                            temp_products.append(temp_product)

                    else:
                        for alist in alists:
                            temp_product = {}
                            products_list = product["products"].copy()
                            products_list.append(f'{alist.find("p").get_text()}')
                            temp_product["product_name"] = product["product_name"]
                            temp_product["products"] = products_list
                            temp_product["path"] = alist["href"]
                            temp_product["finish"] = False
                            temp_products.append(temp_product)

                products = temp_products

            finish = not (all([dic["finish"] for dic in products]))
        return products

    def scroll_page(self, subcategory_url):
        driver = webdriver.Chrome(self.webdriver)
        driver.implicitly_wait(5)

        try:
            SCROLL_PAUSA_TIME = 3
            driver.get(subcategory_url)

            last_height = driver.execute_script("return document.body.scrollHeight")
            while True:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(SCROLL_PAUSA_TIME)
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    time.sleep(3)
                    break
                last_height = new_height
        finally:
            bs = BeautifulSoup(driver.page_source, "html.parser")
            return bs

    def set_products(self, product_category: list):
        products = []
        for data in self.final_data:
            product = Product(
                product_name=data["product_name"],
                products=data["products"],
                path=self.page_address + data["path"],
                page_address=self.page_address,
                products_category=product_category,
            )
            products.append(product.get_product())

        return products

    def set_json(self, products):
        filename = f"files/vw_lcv" + time.strftime("%Y%m%d-%H%M%S") + ".txt"
        with open(filename, "w") as outfile:
            json.dump(products, outfile)

    def set_csv_fb(self, products):
        header = [
            "id",
            "title",
            "description",
            "availability",
            "condition",
            "price",
            "link",
            "image_link",
            "brand",
            "additional_image_link",
            "google_product_category",
        ]

        filename = f"files/vw_lcv_fb_.csv"
        with open(filename, "w", newline="", encoding="utf-8") as output_file:
            thewriter = csv.writer(output_file)
            thewriter.writerow(header)
            for product in products:
                thewriter.writerow(
                    [
                        product["id"],
                        product["sub_name"],
                        product["description"],
                        product["availability"],
                        product["condition"],
                        product["price"],
                        product["link"],
                        product["image_link"],
                        product["brand"],
                        product["additional_image_link"],
                        product["google_product_category"],
                    ]
                )

            original = filename
            target = (
                filename.replace(".csv", "") + time.strftime("%Y%m%d-%H%M%S") + ".csv"
            )
            shutil.copyfile(original, target)

    def set_csv_gmc(self, products):
        header = [
            "id",
            "title",
            "link",
            "price",
            "shipping",
            "image_link",
            "availability",
            "brand",
            "condition",
            "google_product_category",
            "custom_label_0",
            "custom_label_1",
            "ads_label",
            "description",
        ]

        filename = f"files/vw_lcv_gmc_.csv"
        with open(filename, "w", newline="", encoding="utf-8") as output_file:
            thewriter = csv.writer(output_file)
            thewriter.writerow(header)
            for product in products:
                thewriter.writerow(
                    [
                        product["id"],
                        product["sub_name"],
                        product["link"],
                        product["price"],
                        product["sipping"],
                        product["image_link"],
                        product["availability"],
                        product["brand"],
                        product["condition"],
                        product["google_product_category_nr"],
                        product["custom_label_0"],
                        product["availability_"],
                        product["custom_label_0"],
                        product["description"],
                    ]
                )
            original = filename
            target = (
                filename.replace(".csv", "") + time.strftime("%Y%m%d-%H%M%S") + ".csv"
            )
            shutil.copyfile(original, target)

    def remove_duplicate(self, products):
        unic_id = []
        unic_products = []
        removed_id = []

        for product in reversed(products):
            if product["id"] not in unic_id:
                unic_products.append(product)
                unic_id.append(product["id"])
            else:
                removed_id.append(product["id"])

        with open(r'files/removed_id.txt', 'w') as fp:
            for item in removed_id:
                # write each item on a new line
                fp.write("%s\n" % item)
        return unic_products

    def get_tshirt_size(self, products: list, product_category: list) -> list:
        tshirts = []

        for product in products:
            if product["tshirt_links"] is not None:
                for tshirt_link in product["tshirt_links"]:
                    tshirt = Product(
                        product_name=product["main_category"],
                        products=product["subproducts"],
                        path=tshirt_link,
                        page_address=self.page_address,
                        products_category=product_category,
                        tshirt=False,
                    )
                    tshirts.append(tshirt.get_product())

        products.extend(tshirts)
        return products


VW_LCV()
