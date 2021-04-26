from openpyxl import load_workbook

rows = []

def get_data():
    wb = load_workbook('products.xlsx')
    ws = wb.worksheets[0]
    return [row for row in ws.iter_rows(values_only=True)]


class FindProduct:
    def __init__(self, rows):
        self.rows = rows
        self.row = None

    def get_product(self, product: str):
        for row in self.rows:
            if product.upper().strip() in row[0].upper().strip():
                self.row = [x for x in row if x is not None]
                break

    def get_google_nr(self):
        return self.row[1] if self.row is not None else None

    def get_last_category_google(self):
        return self.row[-1] if self.row is not None else None

    def get_fb_product_path(self):
        return ' > '.join(self.row[2:]) if self.row is not None else None

if __name__ == '__main__':
    rows = get_data()
    find_product = FindProduct(rows)
    find_product.get_product('hak')
    nr = find_product.get_google_nr()
    last_category = find_product.get_last_category_google()
    fb_path = find_product.get_fb_product_path()
    pass
